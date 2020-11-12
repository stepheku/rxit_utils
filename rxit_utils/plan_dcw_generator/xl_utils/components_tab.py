'''
components_tab.py
~~~~~~~~~~~~~~~~~~~~
This module contains openpyxl functions related to the Components tab
of the PowerPlan DCW generator
'''

import openpyxl
from openpyxl.styles.fills import GradientFill, Stop, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.fonts import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles.alignment import Alignment
from . import general_utils as util
from . import borders as util_borders
from . import shared_tab_rows as shared_tab_rows

def format_components_tab(workbook: openpyxl.Workbook, 
                          output_file: str=None):
    '''
    Formats the Components tab of the DCW
    '''

    sheet_name = 'Components'

    util.check_toc_tab_exists(workbook=workbook, sheet_name=sheet_name)

    worksheet = workbook[sheet_name]

    # Column dimensions:
    worksheet.column_dimensions['A'].width = 15.85546875
    worksheet.column_dimensions['B'].width = 44.0
    worksheet.column_dimensions['C'].width = 21.28515625
    worksheet.column_dimensions['D'].width = 40
    worksheet.column_dimensions['E'].width = 40
    worksheet.column_dimensions['F'].width = 9.140625
    worksheet.column_dimensions['G'].width = 8.7109375
    worksheet.column_dimensions['H'].width = 8.7109375
    worksheet.column_dimensions['I'].width = 9.42578125
    worksheet.column_dimensions['J'].width = 15.85546875
    worksheet.column_dimensions['K'].width = 15.0
    worksheet.column_dimensions['L'].width = 11.28515625
    worksheet.column_dimensions['M'].width = 9.42578125

    # Row dimensions
    worksheet.row_dimensions[1].height = 33.75
    worksheet.row_dimensions[2].height = 21.75
    worksheet.row_dimensions[3].height = 45

    # Row 1 formatting
    bg_fill = GradientFill(type='linear', degree=90,
                           stop=[
                               Stop(
                                   position=0, color=openpyxl.styles.colors.Color(rgb='FF2A92D0')
                               ),
                               Stop(
                                   position=1, color=openpyxl.styles.colors.Color(rgb='FF0F6EB4')
                               )
                           ])
    bg_font = Font(name='Segoe UI Light', sz=20,
                   family=2.0, b=True, color=Color(rgb='FFFFFFFF'))


    for col in range(1, 17):
        # Non-hyperlinked cells
        active_cell = worksheet.cell(row=1, column=col)
        active_cell.font = bg_font
        active_cell.fill = bg_fill

    worksheet = add_links_header_to_components_tab(worksheet, 
                                        output_file=output_file)

    return workbook


def add_links_header_to_components_tab(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, output_file: str=None):

    worksheet = shared_tab_rows.add_links_header(
        link='Table of Contents', col_start=6, col_end=7, 
        worksheet=worksheet, output_file=output_file)
    worksheet = shared_tab_rows.add_links_header(
        link='Plan Attributes', col_start=8, col_end=9, 
        worksheet=worksheet, output_file=output_file)
    worksheet = shared_tab_rows.add_links_header(
        link='Treatment Schedules', col_start=10, col_end=11, 
        worksheet=worksheet, output_file=output_file)

    return worksheet


def format_components_tab_row_2(workbook: openpyxl.Workbook):
    '''
    Formats row 2 in the Components tab of the DCW
    '''

    sheet_name = 'Components'

    util.check_toc_tab_exists(workbook=workbook, sheet_name=sheet_name)

    worksheet = workbook[sheet_name]

    bg_fill = GradientFill(type='linear', degree=90,
                           stop=[
                               Stop(
                                   position=0, color=openpyxl.styles.colors.Color(rgb='FF344756')
                               ),
                               Stop(
                                   position=1, color=openpyxl.styles.colors.Color(rgb='FF1A2D3B')
                               )
                           ])
    bg_font = Font(name='Calibri', sz=16,
                   family=2.0, b=True, color=Color(rgb='FFFFFFFF'))

    for col in range(1, 17):
        # Non-hyperlinked cells
        active_cell = worksheet.cell(row=2, column=col)
        active_cell.font = bg_font
        active_cell.fill = bg_fill
        if (col == 1):
            util_borders.set_top_bot_l_borders(active_cell)
        elif (col == 16):
            util_borders.set_top_bot_r_borders(active_cell)
        else:
            util_borders.set_top_bot_borders(active_cell)

    return workbook

def format_components_tab_row_3(workbook: openpyxl.Workbook):
    '''
    Formats row 2 in the Components tab of the DCW
    '''

    sheet_name = 'Components'

    util.check_toc_tab_exists(workbook=workbook, sheet_name=sheet_name)

    worksheet = workbook[sheet_name]

    field_bg_fill = PatternFill(patternType='solid', 
                            fgColor=Color(rgb='FF48545D', type='rgb'),
                            bgColor=Color(indexed=64, type='indexed', theme=0))

    field_font = Font(name='Calibri', family=2, sz=11, b=True,
                     scheme='minor', vertAlign=None, color=Color(theme=0, type='theme'))

    for col in range(1, 17):
        active_cell = worksheet.cell(row=3, column=col)
        active_cell.fill = field_bg_fill
        active_cell.font = field_font
        active_cell.alignment = Alignment(wrapText=True)
        util_borders.set_outside_borders(active_cell)

    return workbook

def add_labels_components_tab(workbook: openpyxl.Workbook):
    '''
    Adds labels to the Components tab of the DCW
    '''

    sheet_name = 'Components'

    util.check_toc_tab_exists(workbook=workbook, sheet_name=sheet_name)

    worksheet = workbook[sheet_name]

    worksheet['A1'].value = '=\'Table of Contents\'!B2'
    worksheet['A2'].value = 'PowerPlan Components'


    field_labels = ['Component Type','Component',
        'IV Ingredient (for multi-ingredient IV sets)','Order Sentence Detail',
        'Order Comments','No default sentence','Required ','Prechecked',
        'Time Zero Flag','Time Zero Offset','Offset','Target Duration',
        'Allow Proactive Eval','Chemo','Chemo Related','Persistent Note',]

    for col, val in enumerate(field_labels, 1):
        active_cell = worksheet.cell(row=3, column=col)
        active_cell.value = val

    return workbook
    
def set_up_components_tab(workbook: openpyxl.Workbook,
                          output_file: str=None):
    """
    Sets up the Components tab of the DCW
    """

    format_components_tab(workbook, output_file=output_file)
    add_labels_components_tab(workbook)
    format_components_tab_row_2(workbook)
    format_components_tab_row_3(workbook)

    return workbook

def set_word_wrap_across_columns(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, min_col:int, max_col: int):
    for row in worksheet.iter_cols(min_col=min_col, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(wrapText=True)
    return worksheet


def set_horiz_vert_center_across_columns(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, min_col:int, max_col: int):
    for row in worksheet.iter_cols(min_col=min_col, max_col=max_col, min_row=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    return worksheet