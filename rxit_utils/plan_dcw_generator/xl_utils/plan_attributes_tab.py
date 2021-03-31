"""
plan_attributes_tab.py
~~~~~~~~~~~~~~~~~~~~
This module contains openpyxl functions related to the Plan Attributes
tab of the PowerPlan DCW generator
"""

from collections import OrderedDict
import openpyxl
from openpyxl.styles.fills import GradientFill, Stop, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.fonts import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles.alignment import Alignment
from . import general_utils as util
from . import borders as util_borders
from . import shared_tab_rows as shared_tab_rows
from . import dict_loop as dict_loop


def format_plan_attributes_tab(workbook: openpyxl.Workbook, output_file: str = None):
    """
    Formats the Plan Attributes tab of the DCW
    """

    sheet_name = "Plan Attributes"

    util.check_toc_tab_exists(workbook=workbook, sheet_name=sheet_name)

    worksheet = workbook[sheet_name]

    # Column dimensions:
    worksheet.column_dimensions["A"].width = 43.140625
    worksheet.column_dimensions["B"].width = 26
    worksheet.column_dimensions["C"].width = 13.42578125
    worksheet.column_dimensions["D"].width = 13
    worksheet.column_dimensions["E"].width = 13
    worksheet.column_dimensions["F"].width = 17
    worksheet.column_dimensions["G"].width = 17
    worksheet.column_dimensions["H"].width = 17
    worksheet.column_dimensions["I"].width = 17

    # Row dimensions
    worksheet.row_dimensions[1].height = 33.75

    worksheet = shared_tab_rows.add_worksheet_header(worksheet)
    worksheet = add_links_header_to_plan_attributes_tab(
        worksheet, output_file=output_file
    )

    return workbook


def add_links_header_to_plan_attributes_tab(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, output_file: str = None
):
    """
    Adds links to the header of the Plan Attributes Tab (each tab has their
    own specific links)
    """

    worksheet = shared_tab_rows.add_links_header(
        link="Table of Contents",
        col_start=8,
        col_end=9,
        worksheet=worksheet,
        output_file=output_file,
    )
    worksheet = shared_tab_rows.add_links_header(
        link="Treatment Schedules",
        col_start=10,
        col_end=11,
        worksheet=worksheet,
        output_file=output_file,
    )
    worksheet = shared_tab_rows.add_links_header(
        link="Components",
        col_start=12,
        col_end=13,
        worksheet=worksheet,
        output_file=output_file,
    )

    return worksheet


def add_phase_row(phase_name: str, worksheet: openpyxl.worksheet.worksheet.Worksheet):

    row = util.get_next_empty_row_in_col(col=1, worksheet=worksheet)

    worksheet.row_dimensions[row].height = 21.75

    bg_fill = GradientFill(
        type="linear",
        degree=90,
        stop=[
            Stop(position=0, color=openpyxl.styles.colors.Color(rgb="FF344756")),
            Stop(position=1, color=openpyxl.styles.colors.Color(rgb="FF1A2D3B")),
        ],
    )
    bg_font = Font(
        name="Calibri", sz=16, family=2.0, b=True, color=Color(rgb="FFFFFFFF")
    )

    phase_row_value = "Phase - {}".format(phase_name)

    for col in range(1, 17):
        # Non-hyperlinked cells
        active_cell = worksheet.cell(row=row, column=col)
        active_cell.font = bg_font
        active_cell.fill = bg_fill
        if col == 1:
            util_borders.set_top_bot_l_borders(active_cell)
            active_cell.value = phase_row_value
        elif col == 16:
            util_borders.set_top_bot_r_borders(active_cell)
        else:
            util_borders.set_top_bot_borders(active_cell)

    return worksheet


def add_parent_plan_properties_rows(
    plan_dict: dict, worksheet: openpyxl.worksheet.worksheet.Worksheet
):
    for phase_dict in plan_dict.get("phases").values():
        if phase_dict.get("phase_offset_qty"):
            phase_offset_dict = phase_dict
            break
    if "phase_offset_dict" not in locals():
        phase_offset_dict = {}
    attribute_label_dict = OrderedDict(
        {
            "Plan Name/Display Description": plan_dict.get("display_description"),
            "Description": plan_dict.get("description"),
            "Plan Type": plan_dict.get("plan_type"),
            "Plan Display Method": plan_dict.get("display_method"),
            "Active": "Yes",
            "Version": 1,
            "Begin Effective Date": "",
            "End Effective Date": "",
            "PowerPlan Status": "Production",
            "Cross Encounter": util.bool_to_str(plan_dict.get("cross_encounter_ind")),
            "Evidence Link": plan_dict.get("evidence_link"),
            "Allow Diagnosis Propagation": util.bool_to_str(
                plan_dict.get("allow_diagnosis_propagation_ind")
            ),
            "Hide Flexed Components": util.bool_to_str(
                plan_dict.get("hide_flexed_components")
            ),
            "Use Cycle Numbers": plan_dict.get("use_cycle_numbers_ind"),
            "Standard": "Yes" if plan_dict.get("cycle_std_nbr") else "No",
            "Standard number of cycles": plan_dict.get("cycle_std_nbr"),
            "Range": plan_dict.get(""),
            "Range Begin Value": plan_dict.get(""),
            "Range End Value": plan_dict.get(""),
            "Cycle Increment": plan_dict.get(""),
            "Display Standard Number/End Value": plan_dict.get(""),
            "Restrict ability to modify Standard/End Value": plan_dict.get(""),
            "Cycle Value Display": plan_dict.get("cycle_disp_val"),
            "Prompt for Ordering Physician": util.bool_to_str(
                plan_dict.get("prompt_for_ordering_physician_ind")
            ),
            "Copy Forward": util.bool_to_str(plan_dict.get("copy_forward_ind")),
            "Phase Offsets": plan_dict.get(""),
            "Lab Phase": plan_dict.get(""),
            "Offset": "{} {}".format(phase_offset_dict.get("phase_offset_qty"), phase_offset_dict.get("phase_offset_unit")),
            "Anchor": phase_offset_dict.get("anchor_phase"),
            "Plan Ordering Defaults": plan_dict.get(""),
        }
    )

    row = util.get_next_empty_row_in_col(col=1, worksheet=worksheet)
    for k, v in attribute_label_dict.items():
        active_cell = worksheet.cell(row=row, column=1)
        active_cell.value = k
        active_cell.font = Font(b=True)
        active_cell.offset(column=1).value = v
        row += 1

    return worksheet


def add_plan_ordering_defaults(
    plan_dict: dict, worksheet: openpyxl.worksheet.worksheet.Worksheet
):
    row = util.get_next_empty_row_in_col(col=1, worksheet=worksheet)
    attribute_label_dict = {
        "Prompt user for plan start date and venue": util.bool_to_str(
            plan_dict.get("plan_ord_def_prompt_user_ind")
        ),
        "Open by Default:": plan_dict.get("plan_ord_def_open_default"),
        "Select default visit type": plan_dict.get("plan_ord_def_default_visit"),
    }

    plan_ord_def_columns = [
        "Primary Phase",
        "Optional Phase",
        "Future Phase",
        "This Visit (OP)",
        "This Visit (IP)",
        "Future Visit (OP)",
        "Future Visit (IP)",
    ]

    cell_column = {
        "description": 2,
        "primary_phase_ind": 3,
        "optional_phase_ind": 4,
        "future_phase_ind": 5,
        "this_visit_outpt": 6,
        "this_visit_inpt": 7,
        "future_visit_outpt": 8,
        "future_visit_inpt": 9,
    }

    for k, v in attribute_label_dict.items():
        active_cell = worksheet.cell(row=row, column=1)
        active_cell.value = k
        active_cell.font = Font(b=True)
        active_cell.offset(column=1).value = v
        row += 1

    active_cell = util_borders.set_outside_borders(worksheet.cell(row=row, column=2))

    for idx, val in enumerate(plan_ord_def_columns, start=3):
        active_cell = worksheet.cell(row=row, column=idx)
        active_cell.value = val
        active_cell.font = Font(b=True)
        active_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        active_cell = util_borders.set_outside_borders(active_cell)
    row += 1

    for _, v in sorted(plan_dict["phases"].items(), key=dict_loop.get_phase_seq):
        for k2, v2 in v.items():
            if k2 in cell_column:
                active_cell = worksheet.cell(row=row, column=cell_column.get(k2))
                if isinstance(v2, bool) and v2:
                    active_cell.value = "x"
                    active_cell.alignment = Alignment(horizontal="center")
                elif isinstance(v2, bool) and not v2:
                    active_cell.value = ""
                else:
                    active_cell.value = v2
                active_cell = util_borders.set_outside_borders(active_cell)

        row += 1

    return worksheet


def add_phase_properties(
    phase_dict: dict, worksheet: openpyxl.worksheet.worksheet.Worksheet
):
    attribute_label_dict = OrderedDict(
        {
            "Phase Description": phase_dict.get("description"),
            "Reference Text": phase_dict.get(""),
            "Evidence Link": phase_dict.get(""),
            "Related Results": phase_dict.get(""),
            "Check Alerts on Planning": util.bool_to_str(
                phase_dict.get("check_alerts_on_planning_ind")
            ),
            "Check Alerts on Plan Updates": util.bool_to_str(
                phase_dict.get("check_alerts_on_plan_updt_ind")
            ),
            "Treatment Schedule": "Yes" if phase_dict.get("treatment_sched") else "No",
            "Route for Review": phase_dict.get("route_for_review"),
            # TODO: Add fields here to the query,
            "Classification": phase_dict.get("classification"),
            "Document Reschedule Reason": phase_dict.get(""),
            "Build Linked Components Group": phase_dict.get(""),
        }
    )

    row = util.get_next_empty_row_in_col(col=1, worksheet=worksheet)
    for k, v in attribute_label_dict.items():
        active_cell = worksheet.cell(row=row, column=1)
        active_cell.value = k
        active_cell.font = Font(bold=True)
        active_cell.offset(column=1).value = v
        row += 1

        
    return worksheet


def set_word_wrap_across_columns(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, min_col: int, max_col: int
):
    for row in worksheet.iter_cols(min_col=min_col, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(wrapText=True)
    return worksheet


def set_horiz_vert_center_across_columns(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, min_col: int, max_col: int
):
    for row in worksheet.iter_cols(min_col=min_col, max_col=max_col, min_row=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    return worksheet
