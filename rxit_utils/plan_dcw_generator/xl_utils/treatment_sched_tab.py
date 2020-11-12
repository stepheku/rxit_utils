"""
treatment_sched_tab.py
~~~~~~~~~~~~~~~~~~~~
This module contains openpyxl functions related to the Treatment Schedules 
tab of the PowerPlan DCW generator
"""

import openpyxl
from openpyxl.styles.fills import GradientFill, Stop, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.fonts import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles.alignment import Alignment
from . import general_utils as util
from . import borders as util_borders
from . import shared_tab_rows as shared_tab_rows
from . import dict_loop as dict_loop


def format_treatment_sched_tab(workbook: openpyxl.Workbook, output_file: str = None):
    """
    Formats the Treatment Schedules tab of the DCW
    """

    sheet_name = "Treatment Schedules"

    util.check_toc_tab_exists(workbook=workbook, sheet_name=sheet_name)

    worksheet = workbook[sheet_name]

    # Column dimensions:
    worksheet.column_dimensions["A"].width = 58.42578125
    worksheet.column_dimensions["B"].width = 13
    worksheet.column_dimensions["C"].width = 8.85546875
    worksheet.column_dimensions["D"].width = 13.140625
    worksheet.column_dimensions["E"].width = 17.85546875

    # Row dimensions
    worksheet.row_dimensions[1].height = 33.75

    worksheet = shared_tab_rows.add_worksheet_header(worksheet)
    worksheet = add_links_header_to_treatment_sched_tab(
        worksheet, output_file=output_file
    )

    return workbook


def add_links_header_to_treatment_sched_tab(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, output_file: str = None
):

    worksheet = shared_tab_rows.add_links_header(
        link="Table of Contents",
        col_start=7,
        col_end=8,
        worksheet=worksheet,
        output_file=output_file,
    )
    worksheet = shared_tab_rows.add_links_header(
        link="Plan Attributes",
        col_start=9,
        col_end=10,
        worksheet=worksheet,
        output_file=output_file,
    )
    worksheet = shared_tab_rows.add_links_header(
        link="Components",
        col_start=11,
        col_end=12,
        worksheet=worksheet,
        output_file=output_file,
    )

    return worksheet


def add_phase_row(
    phase_name: str,
    treatment_sched_list: list,
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> openpyxl.worksheet.worksheet.Worksheet:

    row = util.get_next_empty_row_in_col(col=1, worksheet=worksheet)

    worksheet.row_dimensions[row].height = 21.75

    bg_fill = GradientFill(
        type="linear",
        degree=90,
        stop=[
            Stop(position=0, color=openpyxl.styles.colors.Color(rgb="FF344756")),
            Stop(position=1, color=openpyxl.styles.colors.Color(rgb="FF1A2D3B")),
        ],
    )
    bg_font = Font(
        name="Calibri", sz=16, family=2.0, b=True, color=Color(rgb="FFFFFFFF")
    )

    if phase_name == "Scheduling":
        phase_row_value = "Scheduled Phases"
    else:
        phase_row_value = "Treatment Schedule - {}".format(phase_name)

    if len(treatment_sched_list) > 0 or phase_name == "Scheduling":
        for col in range(1, 17):
            # Non-hyperlinked cells
            active_cell = worksheet.cell(row=row, column=col)
            active_cell.font = bg_font
            active_cell.fill = bg_fill
            if col == 1:
                util_borders.set_top_bot_l_borders(active_cell)
                active_cell.value = phase_row_value
            elif col == 16:
                util_borders.set_top_bot_r_borders(active_cell)
            else:
                util_borders.set_top_bot_borders(active_cell)

    return worksheet


def add_phase_labels_and_properties(
    phase_name: str,
    treatment_sched_list: list,
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    duration: int = "",
    duration_unit: str = "",
) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Adds phase-property labels and values (Phase Description, Duration and
    Duration Unit)
    """
    if phase_name != "Scheduling":
        phase_labels = [
            ("Phase Description", phase_name),
            ("Duration", duration),
            ("Duration Unit", duration_unit),
        ]
    else:
        phase_labels = [
            ("Phase Description", phase_name),
        ]

    next_row = util.get_next_empty_row_in_col(col=1, worksheet=worksheet)

    if phase_name == "Scheduling" or len(treatment_sched_list) > 0:
        for row, label_tup in enumerate(phase_labels, next_row):
            active_cell = worksheet.cell(row=row, column=1)
            active_cell.font = Font(name="Calibri", sz=11, family=2.0, b=True)
            active_cell.value = label_tup[0]
            active_cell.offset(column=1).value = label_tup[1]
            active_cell.offset(column=1).alignment = Alignment(horizontal="left")

    return worksheet


def add_treatment_period_labels(
    phase_name: str,
    treatment_sched_list: list,
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Adds the labels for PowerPlan available treatment periods (Treatment Period,
    Primary, Duration, etc)
    """
    field_labels = [
        "Treatment Period",
        "Primary",
        "Duration",
        "Duration Unit",
        "Treatment Interval",
    ]

    field_label_bg_fill = PatternFill(
        patternType="solid",
        fgColor=Color(type="rgb", rgb="FFDBDBDB"),
        bgColor=Color(type="indexed", indexed=64),
    )

    if phase_name != "Scheduling" and len(treatment_sched_list) > 0:
        next_row = util.get_next_empty_row_in_col(col=1, worksheet=worksheet)

        for col, val in enumerate(field_labels, 1):
            active_cell = worksheet.cell(row=next_row, column=col)
            active_cell.font = Font(name="Calibri", sz=11, family=2.0, b=True)
            active_cell.value = val

        for x in range(1, 17):
            active_cell = worksheet.cell(row=next_row, column=x)
            active_cell.fill = field_label_bg_fill

    return worksheet


def add_treatment_periods(
    phase_name: str,
    treatment_sched_list: list,
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
):
    """
    Adds treatment periods to the Treatment Schedules tab
    """
    if phase_name == "Scheduling":
        return worksheet

    if not isinstance(treatment_sched_list, list) or len(treatment_sched_list) < 1:
        return worksheet

    for treatment_period in treatment_sched_list:
        next_row = util.get_next_empty_row_in_col(col=1, worksheet=worksheet)
        active_cell = worksheet.cell(row=next_row, column=1)
        active_cell.value = treatment_period[0]
        active_cell.font = Font(name="Calibri", sz=11, family=2.0, b=True)

        duration_cell = active_cell.offset(column=2)
        duration_cell.value = int(treatment_period[1])

        duration_unit_cell = duration_cell.offset(column=1)
        duration_unit_cell.value = treatment_period[2]

    return worksheet


def add_component_labels(
    phase_name: str,
    treatment_sched_list: list,
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
):
    """
    Add labels for components (Components, followed by the DOTs)
    """
    if len(treatment_sched_list) < 1 and phase_name != "Scheduling":
        return worksheet

    next_row = util.get_next_empty_row_in_col(col=1, worksheet=worksheet)

    field_label_bg_fill = PatternFill(
        patternType="solid",
        fgColor=Color(type="rgb", rgb="FFDBDBDB"),
        bgColor=Color(type="indexed", indexed=64),
    )
    field_label_font = Font(name="Calibri", sz=11, family=2.0, b=True)

    component_label_cell = worksheet.cell(row=next_row, column=1)

    if phase_name == "Scheduling":
        active_cell = worksheet.cell(row=next_row, column=1)
        active_cell.value = "Scheduling Order"
        active_cell.fill = field_label_bg_fill
        active_cell.font = field_label_font
        active_cell.offset(column=1).value = "Link"
        active_cell.offset(column=1).fill = field_label_bg_fill
        active_cell.offset(column=1).font = field_label_font
        return worksheet

    else:
        component_label_cell.value = "Component"

    for idx, treatment_period in enumerate(treatment_sched_list, 1):
        component_label_cell.offset(column=idx).value = treatment_period[0]

    for col in range(1, 17):
        active_cell = worksheet.cell(row=next_row, column=col)
        active_cell.font = field_label_font
        active_cell.fill = field_label_bg_fill
        if col > 1:
            active_cell.alignment = Alignment(horizontal="center")

    return worksheet


def add_components(
    phase_name: str,
    treatment_sched_list: list,
    component_dict: dict,
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
):
    """
    Add labels for components (Components, followed by the DOTs)
    """

    for _, v in sorted(component_dict.items(), key=dict_loop.get_comp_seq):
        if v.get("treatment_sched"):
            next_row = util.get_next_empty_row_in_col(col=1, worksheet=worksheet)
            active_cell = worksheet.cell(row=next_row, column=1)
            active_cell.value = v.get("component")
            for treatment_sched in v.get("treatment_sched"):
                idx = treatment_sched_list.index(treatment_sched)
                active_cell.offset(column=idx + 1).value = "x"
                active_cell.offset(column=idx + 1).alignment = Alignment(
                    horizontal="center"
                )
        elif phase_name == "Scheduling":
            if v.get("component_type") != "Note":
                next_row = util.get_next_empty_row_in_col(col=1, worksheet=worksheet)
                active_cell = worksheet.cell(row=next_row, column=1)
                active_cell.value = v.get("component")
                active_cell.offset(column=1).value = v.get("scheduled_phase")

    return worksheet


def set_word_wrap_across_columns(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, min_col: int, max_col: int
):
    for row in worksheet.iter_cols(min_col=min_col, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(wrapText=True)
    return worksheet


def set_horiz_vert_center_across_columns(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, min_col: int, max_col: int
):
    for row in worksheet.iter_cols(min_col=min_col, max_col=max_col, min_row=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    return worksheet


def add_phase_components_to_treatment_sched_tab(
    phase_name: str,
    treatment_sched_list: list,
    component_dict: dict,
    duration: int,
    duration_unit: str,
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> openpyxl.worksheet.worksheet.Worksheet:

    worksheet = add_phase_row(
        phase_name=phase_name,
        treatment_sched_list=treatment_sched_list,
        worksheet=worksheet,
    )

    worksheet = add_phase_labels_and_properties(
        phase_name=phase_name,
        treatment_sched_list=treatment_sched_list,
        duration=duration,
        duration_unit=duration_unit,
        worksheet=worksheet,
    )

    worksheet = add_treatment_period_labels(
        phase_name=phase_name,
        treatment_sched_list=treatment_sched_list,
        worksheet=worksheet,
    )

    worksheet = add_treatment_periods(
        phase_name=phase_name,
        treatment_sched_list=treatment_sched_list,
        worksheet=worksheet,
    )

    worksheet = add_component_labels(
        phase_name=phase_name,
        treatment_sched_list=treatment_sched_list,
        worksheet=worksheet,
    )

    worksheet = add_components(
        phase_name=phase_name,
        treatment_sched_list=treatment_sched_list,
        component_dict=component_dict,
        worksheet=worksheet,
    )

    return worksheet
