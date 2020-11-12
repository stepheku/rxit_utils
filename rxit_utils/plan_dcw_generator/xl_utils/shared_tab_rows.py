'''
shared_tab_rows.py
~~~~~~~~~~~~~~~~~~~~
This module contains openpyxl functions related to the rows and row types
that are common between the different tabs of a DCW
'''

import openpyxl
from . import borders as util_borders
from openpyxl.styles.fills import GradientFill, Stop, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.fonts import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles.alignment import Alignment


def add_worksheet_header(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    """
    Adds a PowerPlan header for any tab on the PowerPlan DCW. This only adds
    the PowerPlan header and not the links to the different tabs
    """
    bg_fill = GradientFill(type='linear', degree=90,
                           stop=[Stop(position=0, color=Color(rgb='FF2A92D0')),
                                 Stop(position=1, color=Color(rgb='FF0F6EB4'))])
    bg_font = Font(name='Segoe UI Light', sz=20,
                   family=2.0, b=True, color=Color(rgb='FFFFFFFF'))

    for col in range(1, 17):
        active_cell = worksheet.cell(row=1, column=col)
        active_cell.font = bg_font
        active_cell.fill = bg_fill

    # PowerPlan name is already denoted on the Table of Contents tab, so
    # refer to that tab for the name
    worksheet['A1'].value = '=\'Table of Contents\'!B2'

    return worksheet


def add_links_header(link: str, col_start: int, col_end: int,
                     worksheet: openpyxl.worksheet.worksheet.Worksheet,
                     output_file: str = None):
    """
    Adds linked cells to the header for any tab on the PowerPlan DCW. The 
    output_file is needed so that the links will point to the different tabs
    of the same file, and it needs the file name for some reason
    """

    if output_file is None:
        output_file = 'test_output.xlsx'

    link_bg_fill = PatternFill(patternType='solid',
                               fgColor=Color(
                                   tint=-0.25, type='theme', theme=0),
                               bgColor=Color(
                                   indexed=64, type='indexed', theme=0))
    link_font = Font(name='Calibri', family=2, sz=11, u='single',
                     scheme='minor', vertAlign=None,
                     color=Color(theme=10, type='theme'))
    link_alignment = Alignment(horizontal='center', vertical='center',
                               wrap_text=True)

    initial_cell = worksheet.cell(row=1, column=col_start)
    initial_cell.value = link
    initial_cell.hyperlink = '{}#\'{}\'!A1'.format(output_file, link)
    initial_cell.style = 'Hyperlink'

    for x in range(col_start, col_end + 1):
        active_cell = worksheet.cell(row=1, column=x)
        active_cell.font = link_font
        active_cell.alignment = link_alignment
        active_cell.fill = link_bg_fill
        active_cell = util_borders.set_outside_borders(active_cell)

    worksheet.merge_cells(
        start_row=1, start_column=col_start, end_row=1, end_column=col_end)

    return worksheet
