"""
toc_tab.py
~~~~~~~~~~~~~~~~~~~~
This module contains openpyxl functions related to the Table of Contents tab
of the PowerPlan DCW generator
"""

import openpyxl
from . import general_utils as util
from openpyxl.styles.fills import GradientFill, Stop
from openpyxl.styles.colors import Color
from openpyxl.styles.fonts import Font
from openpyxl.worksheet.hyperlink import Hyperlink
import re

def add_text_to_toc_tab(
    workbook: openpyxl.Workbook,
    powerplan_name: str = "",
    powerplan_description: str = "",
    web_link: str = "",
):
    """
    Adds the field names to the table of contents tab
    """

    sheet_name = "Table of Contents"

    util.check_toc_tab_exists(workbook, sheet_name=sheet_name)

    worksheet = workbook[sheet_name]

    field_names = [
        "PowerPlan Name/Display Description",
        "Description",
        "Designer",
        "Design Complete",
        "Builder",
        "Build Complete",
        "Web Link",
        "Components",
        "Treatment Schedules",
        "Plan Attributes",
    ]

    for row, field in enumerate(field_names, start=2):
        worksheet.cell(row=row, column=1).value = field

    link_regex = r"(http.*?pdf$)"
    link_findall = re.findall(link_regex, web_link)
    if link_findall:
        link = link_findall[0]
    else:
        link = ""

    worksheet.cell(row=2, column=2).value = powerplan_name
    worksheet.cell(row=3, column=2).value = powerplan_description
    worksheet.cell(row=4, column=2).value = "Stephen Kung"
    worksheet.cell(row=8, column=2).value = link

    return workbook


def format_toc_tab(workbook: openpyxl.Workbook):
    """
    Formats the table of contents tab for PowerPlan DCWs
    """

    sheet_name = "Table of Contents"
    util.check_toc_tab_exists(workbook, sheet_name=sheet_name)

    worksheet = workbook[sheet_name]

    # Column/row dimensions
    worksheet.column_dimensions["A"].width = 45
    worksheet.column_dimensions["B"].width = 101.42578125
    worksheet.row_dimensions[1].height = 28.5

    # Background color
    bg_fill = GradientFill(
        type="linear",
        degree=90,
        stop=[
            Stop(position=0, color=openpyxl.styles.colors.Color(rgb="FF344756")),
            Stop(position=1, color=openpyxl.styles.colors.Color(rgb="FF1A2D3B")),
        ],
    )
    worksheet["A1"].fill = bg_fill
    worksheet["B1"].fill = bg_fill

    # Font style
    worksheet["A1"].font = Font(
        name="Calibri",
        family=2.0,
        b=True,
        sz=22,
        color=openpyxl.styles.colors.Color(rgb="FFFFFFFF"),
    )
    worksheet["A1"] = sheet_name

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    for row in range(2, 9):
        active_cell = worksheet.cell(row=row, column=1)
        active_cell.font = Font(name="Calibri", sz=11, b=True)

    return workbook


def add_links_on_toc_tab(workbook: openpyxl.Workbook, output_file: str = None):
    """
    Adds Excel links to appropriate cells on the Table of Contents tab
    """
    if output_file is None:
        output_file = "test_output.xlsx"
    sheet_name = "Table of Contents"
    util.check_toc_tab_exists(workbook, sheet_name=sheet_name)

    worksheet = workbook["Table of Contents"]

    linked_cell_values = ["Components", "Treatment Schedules", "Plan Attributes"]

    for row in range(9, 12):
        active_cell = worksheet.cell(row=row, column=1)
        if active_cell.value in linked_cell_values:
            active_cell.hyperlink = "{}#'{}'!A1".format(output_file, active_cell.value)
            # TODO: Hyperlink is dependent on output file name
            active_cell.style = "Hyperlink"
            active_cell.font = Font(
                name="Calibri",
                family=2,
                sz=14,
                u="single",
                color=Color(theme=10, type="theme"),
            )
    return workbook


def setup_toc_tab(
    workbook: openpyxl.Workbook,
    powerplan_name: str = None,
    powerplan_description: str = None,
    web_link: str = None,
    output_file: str = None,
):
    """
    Constructs the workbook and the Table of Contents worksheet
    """
    format_toc_tab(workbook)
    add_text_to_toc_tab(
        workbook,
        powerplan_name=powerplan_name,
        powerplan_description=powerplan_description,
        web_link=web_link,
    )
    add_links_on_toc_tab(workbook, output_file=output_file)
    return workbook