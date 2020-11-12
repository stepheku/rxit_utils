"""
components_tab_rows.py
~~~~~~~~~~~~~~~~~~~~
This module contains openpyxl functions to add rows to the Components tab
of the PowerPlan DCW generator
"""

from openpyxl.styles.fills import GradientFill, Stop, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment
import openpyxl
from . import general_utils as util
from . import dict_loop as dict_loop

def add_ind_to_cell(active_cell: openpyxl.cell.cell.Cell) -> openpyxl.cell.cell.Cell:
    """
    If the attribute is an indicator (such as chemo_ind, persistent note), this
    will add the value and format the cells appropriately
    """
    active_cell.value = "x"
    active_cell.alignment = Alignment(horizontal="center")
    return active_cell


def format_note_cells(
    active_cell: openpyxl.cell.cell.Cell,
    bgcolor_red: int,
    bgcolor_green: int,
    bgcolor_blue: int,
) -> openpyxl.cell.cell.Cell:
    component_type_bg_fill = {
        (177, 226, 186): {
            "pattern_fill": PatternFill(
                patternType="solid",
                fgColor=Color(type="rgb", rgb="FFC6E0B4"),
                bgColor=Color(type="indexed", indexed=64),
            ),
            "component_type": "Note Green",
        },
        (255, 255, 128): {
            "pattern_fill": PatternFill(
                patternType="solid",
                fgColor=Color(type="rgb", rgb="FFFDFD9B"),
                bgColor=Color(type="indexed", indexed=64),
            ),
            "component_type": "Note Yellow",
        },
        (128, 255, 255): {
            "pattern_fill": PatternFill(
                patternType="solid",
                fgColor=Color(type="rgb", rgb="FF00B0F0"),
                bgColor=Color(type="indexed", indexed=64),
            ),
            "component_type": "Note Blue",
        },
        (255, 255, 255): {
            "pattern_fill": PatternFill(fill_type=None),
            "component_type": "Note White",
        },
        (0, 0, 0): {
            "pattern_fill": PatternFill(fill_type=None),
            "component_type": "Note without color",
        },
        (255, 128, 192): {
            "pattern_fill": PatternFill(
                patternType="solid",
                fgColor=Color(type="rgb", rgb="FFFFFFFF"),
                bgColor=Color(type="indexed", indexed=64),
            ),
            "component_type": "Note Pink",
        },
    }

    comp_rgb = (bgcolor_red, bgcolor_green, bgcolor_blue)
    comp_type = component_type_bg_fill.get(comp_rgb).get("component_type")
    comp_fill = component_type_bg_fill.get(comp_rgb).get("pattern_fill")

    active_cell.value = comp_type
    active_cell.alignment = Alignment(wrap_text=True)
    for x in range(0, 16):
        active_cell.offset(column=x).fill = comp_fill

    return active_cell


def add_row_2(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    # comp_row: int = None,
    component_type: str = None,
    component: str = None,
    no_default_order_sentence: bool = None,
    required: bool = None,
    prechecked: bool = None,
    time_zero_ind: bool = None,
    time_zero_offset: str = None,
    offset: str = None,
    bgcolor_red: int = None,
    bgcolor_blue: int = None,
    bgcolor_green: int = None,
    target_duration: str = None,
    dcp_clin_cat: str = None,
    dcp_clin_sub_cat: str = None,
    allow_proactive_eval: bool = None,
    chemo_ind: bool = None,
    chemo_related_ind: bool = None,
    persistent_note: bool = None,
    linking_anchor_comp_ind: bool = None,
    linking_group_desc: str = None,
    linking_rule: str = None,
    linking_rule_quantity: int = None,
    linking_override_reason: str = None,
    order_sentences: dict = None,
):
    cell_column = {
        "component_type": 1,
        "component": 2,
        "iv_ingredient": 3,
        "order_sentence_display_line": 4,
        "order_comment": 5,
        "no_default_order_sentence": 6,
        "required": 7,
        "prechecked": 8,
        "time_zero_ind": 9,
        "time_zero_offset": 10,
        "offset": 11,
        "target_duration": 12,
        "allow_proactive_eval": 13,
        "chemo_ind": 14,
        "chemo_related_ind": 15,
        "persistent_note": 16,
    }

    comp_row = util.get_next_empty_row_in_col(col=1, worksheet=worksheet)

    payload_dict = {
        k: v
        for k, v in locals().items()
        if v is not None
        and v not in ["0", 0]
        and k not in ["worksheet", "comp_row", "cell_column"]
    }

    for k, v in payload_dict.items():
        if k in cell_column:
            comp_col = cell_column.get(k)
            active_cell = worksheet.cell(row=comp_row, column=comp_col)
            if isinstance(v, bool) and v:
                active_cell = add_ind_to_cell(active_cell)
            elif v == "Note":
                active_cell = format_note_cells(
                    active_cell=active_cell,
                    bgcolor_red=bgcolor_red,
                    bgcolor_green=bgcolor_green,
                    bgcolor_blue=bgcolor_blue,
                )
            else:
                if linking_rule_quantity and comp_col == 1:
                    active_cell.value = (
                        "{} - Linked Component Group - {}, {}, {}, {}".format(
                            v,
                            linking_group_desc,
                            linking_rule,
                            linking_rule_quantity,
                            linking_override_reason,
                        )
                    )
                else:
                    active_cell.value = v
                active_cell.alignment = Alignment(wrap_text=True)
    for os_id, os in sorted(
        order_sentences.items(), key=dict_loop.get_order_sentence_seq
    ):
        if os_id != 0:
            sentence_column = cell_column.get("order_sentence_display_line")
            iv_component_column = cell_column.get("iv_ingredient")
            order_comment_column = cell_column.get("order_comment")
            active_cell = worksheet.cell(row=comp_row, column=sentence_column)
            active_cell.value = os.get("order_sentence_display_line")
            active_cell.alignment = Alignment(wrap_text=True)
            active_cell = worksheet.cell(row=comp_row, column=order_comment_column)
            active_cell.value = os.get("order_comment")
            active_cell.alignment = Alignment(wrap_text=True)
            if os.get("iv_ingredient"):
                active_cell = worksheet.cell(row=comp_row, column=iv_component_column)
                active_cell.value = os.get("iv_ingredient")
            comp_row += 1

    return worksheet


def add_multiple_rows(
    comp_dict: dict, worksheet: openpyxl.worksheet.worksheet.Worksheet
):
    """
    Given a dictionary of components and a worksheet, this will add the
    components to the worksheet
    """
    for _, v in sorted(comp_dict.items(), key=dict_loop.get_comp_seq):
        worksheet = add_row_2(
            worksheet=worksheet,
            # comp_row=next_row,
            component_type=v.get("component_type"),
            component=v.get("component"),
            no_default_order_sentence=v.get("no_default_order_sentence"),
            required=v.get("required"),
            prechecked=v.get("prechecked"),
            time_zero_ind=v.get("time_zero_ind"),
            time_zero_offset=v.get("time_zero_offset"),
            offset=v.get("offset"),
            bgcolor_red=v.get("bgcolor_red"),
            bgcolor_blue=v.get("bgcolor_blue"),
            bgcolor_green=v.get("bgcolor_green"),
            target_duration=v.get("target_duration"),
            dcp_clin_cat=v.get("dcp_clin_cat"),
            dcp_clin_sub_cat=v.get("dcp_clin_sub_cat"),
            allow_proactive_eval=v.get("allow_proactive_eval"),
            chemo_ind=v.get("chemo_ind"),
            chemo_related_ind=v.get("chemo_related_ind"),
            persistent_note=v.get("persistent_note"),
            linking_anchor_comp_ind=v.get("linking_anchor_comp_ind"),
            linking_group_desc=v.get("linking_group_desc"),
            linking_rule=v.get("linking_rule"),
            linking_rule_quantity=v.get("linking_rule_quantity"),
            linking_override_reason=v.get("linking_override_reason"),
            order_sentences=v.get("order_sentences"),
        )

    return worksheet


def add_phase_label(
    phase_name: str, row: int, worksheet: openpyxl.worksheet.worksheet.Worksheet
):
    phase_bg_fill = PatternFill(
        patternType="solid",
        fgColor=Color(type="rgb", rgb="FF305496"),
        bgColor=Color(type="indexed", indexed=64),
    )

    phase_font = Font(
        name="Calibri",
        family=2,
        sz=11,
        scheme="minor",
        vertAlign=None,
        color=Color(theme=0, type="theme"),
    )

    worksheet.cell(row=row, column=1).fill = phase_bg_fill
    worksheet.cell(row=row, column=1).font = phase_font
    worksheet.cell(row=row, column=1).value = phase_name

    for x in range(1, 17):
        active_cell = worksheet.cell(row=row, column=x)
        active_cell.fill = phase_bg_fill

    return worksheet