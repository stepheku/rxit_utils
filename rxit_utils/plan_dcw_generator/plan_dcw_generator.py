from pathlib import Path
import argparse
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from .xl_utils import toc_tab as toc_tab
from .xl_utils import general_utils as general_utils
from .xl_utils import components_tab as components_tab
from .xl_utils import components_tab_rows as components_tab_rows
from .xl_utils import dict_loop as dict_loop
from .xl_utils import treatment_sched_tab as treatment_sched_tab
from .xl_utils import plan_attributes_tab as plan_attributes_tab
from . import plan_csv_to_dict


def create_worksheets(workbook: openpyxl.Workbook):
    """
    Creates standard worksheets in the PowerPlan DCWs
    """
    general_utils.check_wb_obj(workbook)

    # Remove any sheets that have the name 'Sheet' (default sheet names)
    for name in workbook.sheetnames:
        if "Sheet" in name:
            workbook.remove(workbook[name])

    new_worksheet_names = [
        "Table of Contents",
        "Components",
        "Treatment Schedules",
        "Plan Attributes",
        "Review questions",
    ]

    for name in new_worksheet_names:
        workbook.create_sheet(title=name)

    return workbook


def replace_illegal_windows_chars(input_str: str) -> str:
    replace_char = "-"
    lookup_dict = {
        "\\": replace_char,
        "/": replace_char,
        "|": replace_char,
        ":": replace_char,
        "<": replace_char,
        ">": replace_char,
        "\"": replace_char,
        "?": replace_char,
    }
    return general_utils.replace_str_by_dict(input_str=input_str, substitutions=lookup_dict)

def main(input_file: str):
    powerplan_dict = plan_csv_to_dict.csv_to_dict(input_file=input_file)
    input_file_path = Path(input_file).parent

    for _, plan_dict in powerplan_dict.items():
        plan_name = plan_dict.get("display_description")
        plan_file_name = plan_name.replace("ONCP ", "")
        plan_file_name = replace_illegal_windows_chars(plan_file_name)

        wb = openpyxl.Workbook()
        output_file = Path(
            input_file_path, "DCW - Oncology PowerPlan {}.xlsx".format(plan_file_name)
        )
        create_worksheets(wb)
        toc_tab.setup_toc_tab(
            workbook=wb,
            powerplan_name=plan_dict["display_description"],
            powerplan_description=plan_dict["description"],
            web_link=plan_dict["evidence_link"],
            output_file=output_file,
        )
        components_tab.set_up_components_tab(workbook=wb, output_file=output_file)
        plan_attributes_tab.format_plan_attributes_tab(
            workbook=wb, output_file=output_file
        )
        treatment_sched_tab.format_treatment_sched_tab(
            workbook=wb, output_file=output_file
        )

        ws = wb["Components"]
        ws_plan_attr = wb["Plan Attributes"]
        ws_treatment_sched = wb["Treatment Schedules"]

        ws_plan_attr = plan_attributes_tab.add_parent_plan_properties_rows(
            plan_dict=plan_dict, worksheet=ws_plan_attr
        )
        ws_plan_attr = plan_attributes_tab.add_plan_ordering_defaults(
            plan_dict=plan_dict, worksheet=ws_plan_attr
        )

        for _, v in sorted(plan_dict["phases"].items(), key=dict_loop.get_phase_seq):
            next_row = general_utils.get_next_empty_row_in_col(col=1, worksheet=ws)
            phase_name = v.get("description")
            comp_dict = v.get("components")
            treatment_sched_list = v.get("treatment_sched")
            duration = v.get("duration_qty")
            duration_unit = v.get("duration_unit")

            phase_duration_dict = {
                "phase": phase_name,
                "duration": duration,
                "duration_unit": duration_unit,
            }

            ws = components_tab_rows.add_phase_label(
                phase_name=phase_name, row=next_row, worksheet=ws
            )

            # Add items to Components tab
            ws = components_tab_rows.add_multiple_rows(
                comp_dict=comp_dict, worksheet=ws
            )

            # Add items to Plan Attributes tab while iterating over phases
            ws_plan_attr = plan_attributes_tab.add_phase_row(
                phase_name=phase_name, worksheet=ws_plan_attr
            )
            ws_plan_attr = plan_attributes_tab.add_phase_properties(
                phase_dict=v, worksheet=ws_plan_attr
            )

            # Add items to Treatment Schedules tab
            ws_treatment_sched = (
                treatment_sched_tab.add_phase_components_to_treatment_sched_tab(
                    phase_name=phase_name,
                    treatment_sched_list=treatment_sched_list,
                    component_dict=comp_dict,
                    duration=duration,
                    duration_unit=duration_unit,
                    worksheet=ws_treatment_sched,
                )
            )

        wb.save(output_file)

if __name__ == "__main__":
    main()